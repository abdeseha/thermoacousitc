from openpyxl import load_workbook
from matplotlib import pyplot as plt
import os 
files = os.listdir()

print('The files in the current folder:')
for index,file in enumerate(files):
    print(str(index)+ ": " + str(file))
file = files[int(input('Enter the nubmer of file to choose: '))]


book = load_workbook(file,data_only=True)
sheets = book.sheetnames

print('The sheets in this file:')
for index,sheet in enumerate(sheets):
     print(str(index) + ': ' + str(sheet))

sheet = book[sheets[int(input("Choose which sheet: "))]]

list = ["P(w)","P_mob_av(w)","P_mob_av2(w)","div(P_mob_av2)"]

for row in (sheet['E2:H2']):
     for index,cell in enumerate(row):
          cell.value = list[index]

i = 3
time = sheet["A" + str(i)].value
volt = 400
p_mob_av = 0
p_mob_av2 = 0
p_list = []
p_list2 = []
while time!= None:
     sheet["E" + str(i)].value = (float(sheet["B" + str(i)].value) + float(sheet["C" + str(i)].value) + float(sheet["D" + str(i)].value))*volt
     p_mob_av += float(sheet["E" + str(i)].value )
     
     if time <= 10:
          p_list.append(float(sheet["E" + str(i)].value))
     
     if time >= 10:
          p_list.append(float(sheet["E" + str(i)].value))
          sheet["F" + str(i)].value = sum(p_list)/len(p_list)
          p_list.pop(0)

     if time >=10 and time <=20:
          p_list2.append(float(sheet["F" + str(i)].value))

     if time >= 20:
          p_list2.append(float(sheet["F" + str(i)].value))
          sheet["G" + str(i)].value = sum(p_list2)/len(p_list2)
          p_list2.pop(0)
          if sheet["G" + str(i-1)].value != None:
               sheet["H" + str(i)].value = (float(sheet["G" + str(i)].value) - float(sheet["G" + str(i-1)].value))/(float(sheet["A" + str(i)].value) - float(sheet["A" + str(i-1)].value))

     i+=1
     time = sheet["A" + str(i)].value

book.save(input("Name of the saving file: ") + ".xlsx")